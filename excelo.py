import openpyxl


'''     CellList(Column, StartRow, EndRow)
"A" |   1   |   3
=> ['A1', 'A2', 'A3']
'''
def CellList(Column, StartRow, EndRow):
    cell_list = [""] * EndRow
    for i_row in range(StartRow - 1,EndRow):
        cell_list[i_row] = Column + str(i_row + 1)
    return cell_list


def GetData(File, Column, Start, End):
    wb = openpyxl.load_workbook(File)
    ws = wb[wb.sheetnames[0]]
    cell_list = CellList( Column, Start, End)
    for i_row in range(Start - 1,End):
        cell_list[i_row] = ws[cell_list[i_row]].value
    wb.close()
    wb.save(File)
    return cell_list


def PutData(File, Column, Start, End, Data):
    wb = openpyxl.load_workbook(File)
    ws = wb[wb.sheetnames[0]]
    cell_list = CellList( Column, Start, End)
    for i_row in range(Start - 1,End):
        ws[cell_list[i_row]].value = Data[i_row]
    wb.close()
    wb.save(File)